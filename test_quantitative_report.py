import processor
import pandas as pd
import tempfile
import os

# Load the extended dataset
df = pd.read_csv('sample_data_extended.csv')

# Create a temporary directory
with tempfile.TemporaryDirectory() as tmpdir:
    print(f'Temporary directory: {tmpdir}')
    
    try:
        # Generate reports
        reports = processor.generate_reports(df, tmpdir, 'Empresa Teste')
        print('Reports generated successfully')
        print('Generated files:', os.listdir(tmpdir))
        
        # Check if Excel file was created
        if 'excel' in reports and os.path.exists(reports['excel']):
            print(f"Excel file created at: {reports['excel']}")
            
            # Load and inspect the Excel file
            import openpyxl
            wb = openpyxl.load_workbook(reports['excel'])
            
            # Check if the new worksheet exists
            if 'Distribuicao_Detalhada' in wb.sheetnames:
                print('New worksheet "Distribuicao_Detalhada" found')
                
                # Display content of the new worksheet
                sheet = wb['Distribuicao_Detalhada']
                print('\nContent of "Distribuicao_Detalhada" worksheet:')
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    print(row)
                    # Limit output for readability
                    if i > 20:
                        print("... (output truncated)")
                        break
            else:
                print('New worksheet "Distribuicao_Detalhada" NOT found')
                
            # List all worksheets
            print("\nAll worksheets in the Excel file:")
            for sheet_name in wb.sheetnames:
                print(f"- {sheet_name}")
        else:
            print("Excel file was not created successfully")
            
    except Exception as e:
        print(f"Error during report generation: {e}")
        import traceback
        traceback.print_exc()