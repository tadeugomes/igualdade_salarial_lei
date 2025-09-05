import processor
import pandas as pd
import tempfile
import os

# Usar o dataset estendido
df = pd.read_csv('sample_data_extended.csv')

# Criar um diretório temporário que não será automaticamente excluído
tmpdir = tempfile.mkdtemp()
print(f'Diretório temporário: {tmpdir}')

try:
    reports = processor.generate_reports(df, tmpdir, 'Empresa Teste')
    print('Arquivos gerados:')
    print(os.listdir(tmpdir))
    
    # Verificar conteúdo da planilha de distribuição por CBO, Sexo e Raça
    import openpyxl
    wb = openpyxl.load_workbook(reports['excel'])
    
    if 'Distribuicao_CBO_Sexo_Raca' in wb.sheetnames:
        sheet = wb['Distribuicao_CBO_Sexo_Raca']
        print('\nConteúdo da planilha Distribuicao_CBO_Sexo_Raca:')
        for row in sheet.iter_rows(values_only=True):
            print(row)
    else:
        print('\nAba Distribuicao_CBO_Sexo_Raca não encontrada')
        
    if 'Distribuicao_CBO' in wb.sheetnames:
        sheet = wb['Distribuicao_CBO']
        print('\nConteúdo da planilha Distribuicao_CBO:')
        for row in sheet.iter_rows(values_only=True):
            print(row)
finally:
    # Não vamos excluir o diretório temporário para que possamos inspecionar os arquivos
    print(f'\nDiretório temporário mantido em: {tmpdir}')